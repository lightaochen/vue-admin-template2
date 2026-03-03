
import os
import sys
import shutil
import re
import time
p = 'TradeAssistant'                                                        # 项目名
root_path = os.path.join(os.path.abspath(__file__).split(p)[0], p)          # 根路径
if root_path not in sys.path:
    sys.path.append(root_path)

from matplotlib import cm
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import platform
import matplotlib.pyplot as plt
from matplotlib import font_manager, colormaps
import matplotlib.gridspec as gridspec
from typing import Optional, Tuple, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from io import BytesIO
from collections import defaultdict
import pytz
from matplotlib.patches import Patch
import colorsys

from models.models import ChinaTradingDay, VChinaFuturesDaybar, ChinaFuturesL1TABar, ChinaFuturesBaseInfo
from common.logger import systemLogger
from common.emailHelper import Mail
from tools.timescaleManager import TimescaleCRUD
from config.settings import *




# 自动选择合适的中文字体
if platform.system() == "Windows":
    font_path = "C:/Windows/Fonts/simhei.ttf"
elif platform.system() == "Darwin":  # macOS
    font_path = "/System/Library/Fonts/STHeiti Medium.ttc"
else:
    font_path = "/usr/share/fonts/truetype/arphic/ukai.ttc"  # Linux 上的中文字体

if os.path.exists(font_path):
    my_font = font_manager.FontProperties(fname=font_path)
    plt.rcParams['font.family'] = my_font.get_name()

plt.rcParams['font.sans-serif'] = ['SimHei', 'WenQuanYi Zen Hei']
plt.rcParams['axes.unicode_minus'] = False


# 定义邮箱服务器和端口
SMTP_SERVER = 'smtp.exmail.qq.com'  # 企业邮箱 SMTP 服务器地址
SMTP_PORT = 465  # SMTP 端口
# 企业邮箱登陆信息
EMAIL_ADDRESS = 'it@aifanzhellc.com'  # 企业微信邮箱地址
EMAIL_PASSWORD = 'QfH9rJ87Bgzm9Ee2'     # 邮箱登录密码
RECEIVERS = [
    "yangsen@avgtech.com.cn",
    "huijiahao@aifanzhellc.com",
    "liyuhao@aifanzhellc.com",
    "yangsen@aifanzhellc.com",
    "janssen.yang@gmail.com",
    "shuojun.li@aifanzhellc.com",
    "yangtao@aifanzhellc.com",
    "wuhua@aifanzhellc.com",
    "zhanggenghuan@aifanzhellc.com",
    "yincai@aifanzhellc.com",
    "panning@aifanzhellc.com",
    "zhaoyi@aifanzhellc.com",
    "jianglang@aifanzhellc.com",
    "haojiashuai@aifanzhellc.com",
    "wangruixuan@aifanzhellc.com",
    "lichao@aifanzhellc.com",
    "hejiaqiang@aifanzhellc.com",
    "liurenyu@aifanzhellc.com",
    "qiantu@aifanzhellc.com",
    "zy_3128@163.com",
    "ranyuzhen@aifanzhellc.com",
    "tianhao@aifanzhellc.com",
    "fuguanhua@aifanzhellc.com",
    "chenguangtao@aifanzhellc.com",
    'zhoujinghao@aifanzhellc.com',
    'huyuyao@aifanzhellc.com'
    ]

# RECEIVERS = [
#     "chenguangtao@aifanzhellc.com"
#     ]



def get_trading_days_in_range(start_day: int, end_day: int, timescale_crud: TimescaleCRUD) -> list:
    """
    使用 TimescaleCRUD 查询指定日期范围内的所有交易日。

    参数:
        start_day (int): 起始日期，格式如 20250704
        end_day (int): 结束日期，格式如 20250804
        timescale_crud (TimescaleCRUD): 数据库操作对象

    返回:
        List[str]: 所有在该区间内的交易日（'YYYY-MM-DD' 格式）
    """
    start_day_dt = datetime.strptime(str(start_day), "%Y%m%d").date()
    end_day_dt = datetime.strptime(str(end_day), "%Y%m%d").date()

    try:
        with timescale_crud.session_scope() as session:
            result = session.query(ChinaTradingDay.trading_day)\
                .filter(ChinaTradingDay.trading_day.between(start_day_dt, end_day_dt))\
                .order_by(ChinaTradingDay.trading_day)\
                .all()
            trading_days = [row.trading_day.strftime("%Y-%m-%d") for row in result]

            systemLogger.info(f"查询交易日成功: 共 {len(trading_days)} 个交易日区间 [{start_day_dt}, {end_day_dt}]")
            return trading_days

    except Exception as e:
        systemLogger.error(f"查询交易日失败: {e}")
        return []

def normalize_trading_day(tradingday):
    """
    将 tradingday 统一转换为 'YYYY-MM-DD' 字符串格式。

    参数支持:
    - datetime 类型；
    - 字符串 'YYYY-MM-DD'；
    - 整数或字符串 'YYYYMMDD'；

    返回:
        str: 标准化后的日期字符串 'YYYY-MM-DD'
    """
    try:
        if isinstance(tradingday, datetime):
            normalized = tradingday.strftime("%Y-%m-%d")
        elif isinstance(tradingday, str) and len(tradingday) == 10 and "-" in tradingday:
            normalized = tradingday  # 已是 'YYYY-MM-DD'
        else:
            normalized = datetime.strptime(str(tradingday), "%Y%m%d").strftime("%Y-%m-%d")

        systemLogger.debug(f"交易日标准化成功: 输入={tradingday}, 输出={normalized}")
        return normalized

    except ValueError as e:
        systemLogger.error(f"交易日标准化失败: 输入={tradingday}, 错误={e}")
        raise ValueError(f"无法识别的交易日格式: {tradingday}")

def get_main_and_sub_contract(
    product: str,
    tradingday: str,
    timescale_crud: TimescaleCRUD,
    previous: bool = False
) -> Tuple[Optional[dict], Optional[dict]]:
    """
    获取某品种在指定交易日的主力与次主力合约信息。

    参数:
        product (str): 品种代码（如 cu, zn）
        tradingday (str): 'YYYY-MM-DD' 格式
        timescale_crud (TimescaleCRUD): 数据库接口
        previous (bool): 是否获取前一交易日的主/次主力

    返回:
        (main_contract_dict, sub_contract_dict)
    """
    try:
        trade_date = datetime.strptime(tradingday, "%Y-%m-%d").date()

        with timescale_crud.session_scope() as session:
            if previous:
                rows = session.query(VChinaFuturesDaybar)\
                    .filter(VChinaFuturesDaybar.product.ilike(product))\
                    .filter(VChinaFuturesDaybar.trading_day < trade_date)\
                    .filter(VChinaFuturesDaybar.rank_volume.in_([1, 2]))\
                    .order_by(VChinaFuturesDaybar.trading_day.desc())\
                    .limit(100)\
                    .all()

                if not rows:
                    systemLogger.warning(f"[前日合约] 查询为空: product={product}, date<{tradingday}")
                    return None, None

                # 找最近一个交易日
                latest_day = max(r.trading_day for r in rows)
                rows = [r for r in rows if r.trading_day == latest_day]

            else:
                rows = session.query(VChinaFuturesDaybar)\
                    .filter(VChinaFuturesDaybar.trading_day == trade_date)\
                    .filter(VChinaFuturesDaybar.product.ilike(product))\
                    .filter(VChinaFuturesDaybar.rank_volume.in_([1, 2]))\
                    .all()

                if not rows:
                    systemLogger.warning(f"[当日合约] 查询为空: product={product}, date={tradingday}")
                    return None, None

            main_contract = sub_contract = None
            for row in rows:
                info = {
                    "instrument_id": row.instrument_id,
                    "high_limited": float(row.high_limited) if row.high_limited is not None else None,
                    "low_limited": float(row.low_limited) if row.low_limited is not None else None,
                    "trading_day": row.trading_day.strftime("%Y-%m-%d")
                }
                if row.rank_volume == 1:
                    main_contract = info
                elif row.rank_volume == 2:
                    sub_contract = info

            systemLogger.info(
                f"主次合约获取成功: product={product}, date={'前一日' if previous else tradingday}, main={main_contract and main_contract['instrument_id']}, sub={sub_contract and sub_contract['instrument_id']}"
            )
            return main_contract, sub_contract

    except Exception as e:
        systemLogger.exception(f"❌ 获取主次合约失败: product={product}, date={tradingday}, previous={previous}, 错误: {e}")
        return None, None



def load_contract_data(instrument_id: str, trading_day: str, timescale_crud, previous: bool = False) -> pd.DataFrame:
    try:
        sh_tz = pytz.timezone("Asia/Shanghai")
        target_day = datetime.strptime(trading_day, "%Y-%m-%d").date()

        with timescale_crud.session_scope() as session:
            if previous:
                pre_day_obj = session.query(ChinaTradingDay.pre_trading_day)\
                    .filter(ChinaTradingDay.trading_day == target_day).first()
                if not pre_day_obj:
                    systemLogger.warning(f"❌ 未找到前一交易日: 当前交易日={trading_day}")
                    return pd.DataFrame()
                target_day = pre_day_obj.pre_trading_day
                systemLogger.info(f"使用前一交易日: 原始={trading_day}, 前一日={target_day}")

            # 更宽松的时间范围，以容纳夜盘数据
            start_dt = datetime.combine(target_day - timedelta(days=14), datetime.min.time()).astimezone(pytz.utc)
            end_dt   = datetime.combine(target_day + timedelta(days=14), datetime.min.time()).astimezone(pytz.utc)



            rows = session.query(ChinaFuturesL1TABar)\
                .filter(ChinaFuturesL1TABar.time.between(start_dt, end_dt))\
                .filter(ChinaFuturesL1TABar.trading_day == target_day)\
                .filter(ChinaFuturesL1TABar.instrument_id.ilike(instrument_id))\
                .order_by(ChinaFuturesL1TABar.time)\
                .all()

            if not rows:
                systemLogger.warning(f"⚠️ 查询分钟数据为空: 合约={instrument_id}, 日期范围={start_dt}~{end_dt}")
                return pd.DataFrame()

            df = pd.DataFrame([{
                **{col.name: getattr(row, col.name) for col in row.__table__.columns}
            } for row in rows])

            if 'time' in df.columns:
                df['time'] = pd.to_datetime(df['time']).dt.tz_convert(sh_tz)

            systemLogger.info(f"✅ 加载分钟数据成功: 合约={instrument_id}, 行数={len(df)}, 日期={target_day}")
            return df

    except Exception as e:
        systemLogger.exception(f"❌ 数据库查询失败: 合约={instrument_id}, 日期={trading_day}, previous={previous}, 错误: {e}")
        return pd.DataFrame()

def load_main_sub_data(timescale_crud: TimescaleCRUD, tradingday: str, product_prefix: str):
    """
    加载某品种在指定交易日的主力与次主力合约数据，并拼接前一交易日尾部数据。

    返回:
        main_df, sub_df: DataFrame（含拼接前一天最后几条记录）
    """

    try:
        systemLogger.info(f"加载主次主力合约: 品种={product_prefix}, 交易日={tradingday}")
        main_info, sub_info = get_main_and_sub_contract(product_prefix, tradingday, timescale_crud)

        if not main_info or not sub_info:
            systemLogger.warning(f"主/次主力合约未找到: 品种={product_prefix}, 交易日={tradingday}")
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

        main_contract = main_info["instrument_id"]
        sub_contract = sub_info["instrument_id"]
        systemLogger.info(f"主力合约={main_contract}, 次主力合约={sub_contract}")

        # 加载当天数据
        main_df = load_contract_data(main_contract, tradingday, timescale_crud)
        main_df["high_limited"] = main_info["high_limited"]
        main_df["low_limited"] = main_info["low_limited"]

        sub_df = load_contract_data(sub_contract, tradingday, timescale_crud)
        sub_df["high_limited"] = sub_info["high_limited"]
        sub_df["low_limited"] = sub_info["low_limited"]

        # 加载前一交易日的数据
        pre_main_df = load_contract_data(main_contract, tradingday, timescale_crud, previous=True)
        pre_sub_df = load_contract_data(sub_contract, tradingday, timescale_crud, previous=True)

        # 拼接前一日尾部数据（通常为夜盘收尾）
        last_main_row = pre_main_df[pre_main_df["instrument_id"] == main_contract].tail(6).copy()
        main_df = pd.concat([last_main_row, main_df], ignore_index=True)

        last_sub_row = pre_sub_df[pre_sub_df["instrument_id"] == sub_contract].tail(6).copy()
        sub_df = pd.concat([last_sub_row, sub_df], ignore_index=True)

        systemLogger.info(
            f"数据加载成功: 品种={product_prefix}, 主力行数={len(main_df)}, 次主力行数={len(sub_df)}"
        )

        return main_df, sub_df, main_info, sub_info

    except Exception as e:
        systemLogger.exception(
            f"❌ 加载主次主力合约数据失败: 品种={product_prefix}, 交易日={tradingday}, 错误={e}"
        )
        raise

def load_pair_data(ts_crud: TimescaleCRUD, tradingday: str, near_ctp: str, far_ctp: str):
    import pandas as pd
    try:
        systemLogger.info(f"手动合约加载: near={near_ctp}, far={far_ctp}, day={tradingday}")

        # 小工具：把 ORM 行转 dict（会话未关闭时调用）
        def row_to_info(row):
            if row is None:
                return None
            return {
                "instrument_id": row.instrument_id,
                "high_limited": float(row.high_limited) if row.high_limited is not None else None,
                "low_limited":  float(row.low_limited)  if row.low_limited  is not None else None,
                "trading_day":  row.trading_day.strftime("%Y-%m-%d") if row.trading_day else None,
            }

        # === 取当日涨跌停（日线）===
        with ts_crud.session_scope() as session:
            trade_date = pd.to_datetime(tradingday).date()
            near_row = (session.query(VChinaFuturesDaybar)
                        .filter(VChinaFuturesDaybar.trading_day == trade_date,
                                VChinaFuturesDaybar.instrument_id.ilike(near_ctp))
                        .first())
            far_row  = (session.query(VChinaFuturesDaybar)
                        .filter(VChinaFuturesDaybar.trading_day == trade_date,
                                VChinaFuturesDaybar.instrument_id.ilike(far_ctp))
                        .first())

            # FIX: 在会话关闭之前就把行对象转成 dict
            near_info = row_to_info(near_row)     # ← dict / None
            far_info  = row_to_info(far_row)      # ← dict / None

            # FIX: 之后用 dict 取值，不再访问 ORM 属性
            near_vals = ({"high_limited": near_info["high_limited"], "low_limited": near_info["low_limited"]}
                         if near_info else None)
            far_vals  = ({"high_limited": far_info["high_limited"], "low_limited": far_info["low_limited"]}
                         if far_info else None)

        # === 加载当天分钟 ===
        near_df = load_contract_data(near_ctp, tradingday, ts_crud)
        far_df  = load_contract_data(far_ctp,  tradingday, ts_crud)

        # 写入涨跌停
        if near_df is not None and not near_df.empty:
            near_df["high_limited"] = near_vals["high_limited"] if near_vals else None
            near_df["low_limited"]  = near_vals["low_limited"]  if near_vals else None
        if far_df is not None and not far_df.empty:
            far_df["high_limited"] = far_vals["high_limited"] if far_vals else None
            far_df["low_limited"]  = far_vals["low_limited"]  if far_vals else None

        # === 前一交易日（尾 6 条）拼接 ===
        pre_near_df = load_contract_data(near_ctp, tradingday, ts_crud, previous=True)
        pre_far_df  = load_contract_data(far_ctp,  tradingday, ts_crud, previous=True)

        if pre_near_df is not None and not pre_near_df.empty:
            near_tail = pre_near_df[pre_near_df["instrument_id"].str.lower() == near_ctp.lower()].tail(6).copy()
            near_df = pd.concat([near_tail, near_df], ignore_index=True)
        if pre_far_df is not None and not pre_far_df.empty:
            far_tail = pre_far_df[pre_far_df["instrument_id"].str.lower() == far_ctp.lower()].tail(6).copy()
            far_df = pd.concat([far_tail, far_df], ignore_index=True)

        systemLogger.info(f"手动合约加载成功: near={near_ctp}({0 if near_df is None else len(near_df)}), "
                          f"far={far_ctp}({0 if far_df is None else len(far_df)})")
        systemLogger.info(f"手动合约当日信息: near={near_info}, far={far_info}, day={tradingday}")

        return near_df, far_df, near_info, far_info

    except Exception as e:
        systemLogger.exception(f"❌ 手动合约加载失败: near={near_ctp}, far={far_ctp}, day={tradingday}, 错误={e}")
        import pandas as pd
        return pd.DataFrame(), pd.DataFrame(), None, None



def get_tick_size(contract_id: str, timescale_crud) -> float:
    """
    获取某合约的最小变动单位（tick_size）

    参数:
        contract_id (str): 合约代码
        timescale_crud (TimescaleCRUD): 数据库接口对象

    返回:
        float or None: tick_size 数值，若未找到则返回 None
    """
    try:
        with timescale_crud.session_scope() as session:
            result = session.query(ChinaFuturesBaseInfo.tick_size)\
                .filter(ChinaFuturesBaseInfo.instrument_id.ilike(contract_id))\
                .limit(1)\
                .first()

            if result and result.tick_size is not None:
                tick = float(result.tick_size)
                systemLogger.info(f"✅ Tick Size 获取成功: 合约={contract_id}, tick_size={tick}")
                return tick
            else:
                systemLogger.warning(f"⚠️ Tick Size 未找到: 合约={contract_id}")
                return None

    except Exception as e:
        systemLogger.exception(f"❌ Tick Size 查询失败: 合约={contract_id}, 错误: {e}")
        return None




def get_avg_volume(timescale_crud, instrument_id: str, end_date: str, days: int = 30) -> int:
    """
    计算某合约过去 N 天的平均成交量（以每个交易日的 volume 汇总为单位）。

    参数:
        timescale_crud (TimescaleCRUD): 数据库接口对象
        instrument_id (str): 合约代码
        end_date (str): 结束日期（格式: YYYY-MM-DD）
        days (int): 统计天数，默认30

    返回:
        int: 平均每日成交量（volume）
    """
    try:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
        start_dt = end_dt - timedelta(days=days)

        start_dt_time = datetime.combine(start_dt - timedelta(days=14), datetime.min.time()).astimezone(pytz.utc)
        end_dt_time   = datetime.combine(end_dt + timedelta(days=14), datetime.min.time()).astimezone(pytz.utc)


        with timescale_crud.session_scope() as session:
            rows = session.query(
                ChinaFuturesL1TABar.trading_day,
                ChinaFuturesL1TABar.volume
            ).filter(
                ChinaFuturesL1TABar.time.between(start_dt_time, end_dt_time),
                ChinaFuturesL1TABar.trading_day.between(start_dt, end_dt),
                ChinaFuturesL1TABar.instrument_id.ilike(instrument_id)

            ).all()

            if not rows:
                systemLogger.warning(f"⚠️ 平均成交量查询为空: 合约={instrument_id}, 区间={start_dt}~{end_dt}")
                return 0

            # 构造 DataFrame 聚合每日 volume 总和
            df = pd.DataFrame(rows, columns=["trading_day", "volume"])
            daily_sum = df.groupby("trading_day")["volume"].sum()

            avg_volume = int(daily_sum.mean())
            systemLogger.info(
                f"✅ 平均成交量查询成功: 合约={instrument_id}, 区间={start_dt}~{end_dt}, 平均={avg_volume}, 有效天数={len(daily_sum)}"
            )
            return avg_volume

    except Exception as e:
        systemLogger.exception(f"❌ 平均成交量查询失败: 合约={instrument_id}, 结束日={end_date}, 错误: {e}")
        return 0



def get_product_prefix(instrument_id: str) -> str:
    # 逐字符遍历，找到第一个非字母位置，然后切片
    for i, ch in enumerate(instrument_id):
        if not ch.isalpha():
            return instrument_id[:i].lower()
    return instrument_id.lower()  # 如果全是字母，返回全部

def remove_open_close_noise(df: pd.DataFrame) -> pd.DataFrame:
    """
    根据主力合约品种规则，剔除不在交易时间段内的数据。

    输入:
        df: 包含 'instrument_id' 和 'time' 列的 DataFrame

    输出:
        过滤后的 DataFrame，移除非交易时间数据
    """
    if df.empty:
        systemLogger.warning("⚠️ 输入数据为空，跳过交易时段过滤")
        return df

    try:
        df["__time_only__"] = pd.to_datetime(df["time"], errors="coerce").dt.time
        df["__product__"] = df["instrument_id"].apply(get_product_prefix)

        def is_in_valid_session(row):
            product = row["__product__"]
            t = row["__time_only__"]
            if pd.isnull(t): return False
            sessions = TRADING_SESSION_RULES.get(product, [])
            for s in sessions:
                start = datetime.strptime(s[0], "%H:%M").time()
                end = datetime.strptime(s[1], "%H:%M").time()
                if start <= end:
                    if start <= t <= end:
                        return True
                else:
                    if t >= start or t <= end:
                        return True
            return False

        before_rows = len(df)
        df = df[df.apply(is_in_valid_session, axis=1)].copy()
        after_rows = len(df)

        systemLogger.info(f"🧹 交易时段过滤: 原始={before_rows} 行，保留={after_rows} 行，过滤={before_rows - after_rows} 行")
        return df.drop(columns=["__time_only__", "__product__"])

    except Exception as e:
        systemLogger.exception(f"❌ 交易时段过滤失败: 错误={e}")
        return df

def align_main_sub_minute(main_df: pd.DataFrame, sub_df: pd.DataFrame, tick_size: float) -> pd.DataFrame:
    """
    对齐主力与次主力分钟数据，并计算中间价、价差（spread）。

    参数:
        main_df: 主力合约数据，含 time, bid/ask_price0
        sub_df: 次主力合约数据
        tick_size: 最小跳动单位

    返回:
        pd.DataFrame: 包含 mid_main, mid_sub, spread 的对齐结果
    """
    try:
        time_main = set(main_df["time"])
        time_sub = set(sub_df["time"])
        common_times = sorted(list(time_main & time_sub))
        all_times = sorted(list(time_main | time_sub))

        systemLogger.info(
            f"⌛ 时间戳对齐检查: 主力={len(time_main)} 次主力={len(time_sub)} 共有={len(common_times)} 总计={len(all_times)}"
        )

        if len(common_times) == 0:
            systemLogger.warning("⚠️ 主力与次主力时间戳无交集，无法对齐")
            return pd.DataFrame()

        # 合并数据（仅保留交集）
        df = pd.merge(main_df, sub_df, on="time", suffixes=("_main", "_sub"))
        systemLogger.info(f"✅ 合并成功: 行数={len(df)}")

        # 中间价计算（强制转 float）
        df["mid_main"] = ((df["bid_price0_main"].astype(float) + df["ask_price0_main"].astype(float)) / 2)
        df["mid_sub"] = ((df["bid_price0_sub"].astype(float) + df["ask_price0_sub"].astype(float)) / 2)

        # Spread 计算（使用 float tick_size）
        raw_spread = df["mid_main"] - df["mid_sub"]
        df["spread"] = np.ceil(raw_spread / float(tick_size)) * float(tick_size)
        # 如果不用ceil，可以替换为：
        # df["spread"] = raw_spread

        systemLogger.debug(
            f"价差样本: spread.head() = {df['spread'].head().tolist()}"
        )

        return df

    except Exception as e:
        systemLogger.exception(f"❌ 对齐分钟数据失败: 错误={e}")
        return pd.DataFrame()

def filter_spread_with_limit(df: pd.DataFrame) -> Tuple[pd.DataFrame, list]:
    """
    若某一分钟主力或次主力触及涨跌停，则该分钟的 spread 设置为 NaN，但 mid 值不变。

    添加两列：
        - main_limited_flag
        - sub_limited_flag

    返回:
        df: 带标记和过滤后的 DataFrame
        limit_timestamps: 触及涨跌停的时间列表（datetime 格式）
    """
    try:
        # 检查是否有涨跌停字段
        for col in [
            "bid_price0_main", "ask_price0_main", "high_limited_main", "low_limited_main",
            "bid_price0_sub", "ask_price0_sub", "high_limited_sub", "low_limited_sub"
        ]:
            if col not in df.columns:
                raise KeyError(f"字段缺失: {col}")

        # 判断是否触及涨跌停
        is_limit_main = (
            (df["bid_price0_main"] >= df["high_limited_main"]) |
            (df["ask_price0_main"] <= df["low_limited_main"])
        )
        is_limit_sub = (
            (df["bid_price0_sub"] >= df["high_limited_sub"]) |
            (df["ask_price0_sub"] <= df["low_limited_sub"])
        )

        # 标记
        df["main_limited_flag"] = is_limit_main
        df["sub_limited_flag"] = is_limit_sub

        # 标记触发时间
        limit_mask = is_limit_main | is_limit_sub
        limit_timestamps = df.loc[limit_mask, "time"].tolist()

        # 屏蔽该分钟的 spread 值
        df.loc[limit_mask, "spread"] = float("nan")

        systemLogger.info(
            f"🚫 涨跌停屏蔽: 主力={is_limit_main.sum()} 分钟, 次主力={is_limit_sub.sum()} 分钟, 总计={len(limit_timestamps)}"
        )
        return df, limit_timestamps

    except Exception as e:
        systemLogger.exception(f"❌ 涨跌停过滤失败: 错误={e}")
        return df, []

def compute_leg_return_series(df: pd.DataFrame) -> pd.Series:
    """
    计算合约中间价序列的分钟变化值（差分序列）
    返回: pd.Series, index = time, dtype=float64
    """
    try:
        if df.empty or "bid_price0" not in df or "ask_price0" not in df or "time" not in df:
            raise ValueError("输入数据缺少必要字段或为空")

        df = df.copy()
        df.set_index("time", inplace=True)

        # 关键：把 Decimal/object 转成 float
        bid = pd.to_numeric(df["bid_price0"], errors="coerce").astype(float)
        ask = pd.to_numeric(df["ask_price0"], errors="coerce").astype(float)

        mid_price = (bid + ask) / 2.0
        diff_series = mid_price.diff()  # 第一条会是 NaN 正常

        return diff_series

    except Exception as e:
        systemLogger.exception(f"❌ compute_leg_return_series 计算失败: 错误={e}")
        return pd.Series(dtype="float64")

def compute_leg_volatility(leg_diff: pd.Series) -> float:
    """
    计算单边腿价格变化序列的波动率（标准差）

    参数:
        leg_diff: 差分序列

    返回:
        float: 标准差值
    """
    try:
        cleaned = leg_diff.dropna().iloc[1:].astype(float)
        if cleaned.empty:
            systemLogger.warning("⚠️ 差分序列为空，无法计算波动率")
            return 0.0

        std_val = cleaned.std(ddof=1)
        systemLogger.info(f"📉 单腿波动率计算: 波动率={std_val:.6f}, 样本数={len(cleaned)}")
        return std_val

    except Exception as e:
        systemLogger.exception(f"❌ compute_leg_volatility 计算失败: 错误={e}")
        return 0.0

def compute_spread_volatility(spread_diff: pd.Series) -> float:
    """
    计算价差变动序列的波动率（样本标准差），跳过第一条（跨日差值）

    参数:
        spread_diff: 差分后的价差序列

    返回:
        float: 波动率（标准差）
    """
    try:
        cleaned = spread_diff.dropna().iloc[1:]
        if cleaned.empty:
            systemLogger.warning("⚠️ 价差差分序列为空，无法计算波动率")
            return 0.0

        std_val = cleaned.std(ddof=1)
        systemLogger.info(f"📊 价差波动率计算成功: 波动率={std_val:.6f}, 样本数={len(cleaned)}")
        return std_val

    except Exception as e:
        systemLogger.exception(f"❌ compute_spread_volatility 计算失败: 错误={e}")
        return 0.0


def compute_spread_leg_correlation(
    spread_diff: pd.Series,
    leg1_diff: pd.Series,
    leg2_diff: pd.Series
) -> dict:
    """
    价差与腿1/腿2、腿1与腿2的相关性；对齐索引，少于2个样本返回0
    """
    try:
        s  = pd.to_numeric(spread_diff, errors="coerce")
        l1 = pd.to_numeric(leg1_diff,    errors="coerce")
        l2 = pd.to_numeric(leg2_diff,    errors="coerce")

        def pair_corr(a: pd.Series, b: pd.Series) -> float:
            df = pd.concat([a, b], axis=1, join="inner").dropna()
            if df.shape[0] < 2:
                return 0.0
            c = df.iloc[:, 0].corr(df.iloc[:, 1])
            try:
                from math import isfinite
                return float(c) if (c is not None and isfinite(float(c))) else 0.0
            except Exception:
                return 0.0

        return {
            "spread_vs_leg1": pair_corr(s,  l1),
            "spread_vs_leg2": pair_corr(s,  l2),
            "leg1_vs_leg2":   pair_corr(l1, l2),
        }

    except Exception as e:
        systemLogger.exception(f"❌ compute_spread_leg_correlation 计算失败(稳健版): 错误={e}")
        return {"spread_vs_leg1": 0.0, "spread_vs_leg2": 0.0, "leg1_vs_leg2": 0.0}

def save_spread_data_to_excel(
    product_prefix: str,
    main_contract: str,
    sub_contract: str,
    stats: dict,
    excel_path: str = "spread_analysis.xlsx"
) -> bool:
    """
    保存主力/次主力/价差分析指标到 Excel 文件中（每个品种一个 sheet，自动追加）。
    价差行按“近月-远月”展示；若主力为远月，则“价格变化量”取相反数。
    """
    try:
        # 近月-远月判定（只影响价差行展示）
        main_is_near, near_c, far_c, sign= _is_main_near(main_contract, sub_contract)
        if main_is_near:
            systemLogger.info(f"Excel展示方向: {near_c}-{far_c}（主力为近月，价差不取反）")
        else:
            systemLogger.info(f"Excel展示方向: {near_c}-{far_c}（主力为远月，价差展示取反）")

        # 展示口径下的价差总变化（取反与否）
        spread_total_change_disp = (stats.get('spread_total_change', 0) or 0) * sign

        sheet_name = product_prefix.upper()

        data = [
            {
                "交易日": stats.get('trading_day', 0),
                "合约类型": "主力合约",
                "合约代码": main_contract,
                "成交量": f"{stats.get('vol_main', 0)}({stats.get('vol_main', 0)/max(1, stats.get('main_avg_month_volume', 1)):.1%})",
                "月均成交量": stats.get('main_avg_month_volume', 0),
                "价格变化量": f"{stats.get('leg1_total_change', 0):.2f}",
                "波动率": f"{stats.get('leg1_volatility', 0):.3f}",
                "与价差相关性": f"{stats.get('corr_spread_leg1', 0):.3f}",
                "涨跌停": "是" if stats.get('main_limit_flag', False) else ""
            },
            {
                "交易日": stats.get('trading_day', 0),
                "合约类型": "次主力合约",
                "合约代码": sub_contract,
                "成交量": f"{stats.get('vol_sub', 0)}({stats.get('vol_sub', 0)/max(1, stats.get('sub_avg_month_volume', 1)):.1%})",
                "月均成交量": stats.get('sub_avg_month_volume', 0),
                "价格变化量": f"{stats.get('leg2_total_change', 0):.2f}",
                "波动率": f"{stats.get('leg2_volatility', 0):.3f}",
                "与价差相关性": f"{stats.get('corr_spread_leg2', 0):.3f}",
                "涨跌停": "是" if stats.get('sub_limit_flag', False) else ""
            },
            {
                "交易日": stats.get('trading_day', 0),
                "合约类型": "价差",
                "合约代码": f"{near_c}-{far_c}",
                "成交量": None,
                "月均成交量": None,
                "价格变化量": f"{spread_total_change_disp:.2f}",               # 注意：展示取反
                "波动率": f"{stats.get('spread_volatility', 0):.3f}",          # 波动率无需取反
                "与价差相关性": f"{stats.get('corr_leg1_leg2', 0):.3f}",
                "涨跌停": None
            }
        ]

        df = pd.DataFrame(data)

        # 确保目录存在
        os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)

        file_exists = os.path.exists(excel_path)
        with pd.ExcelWriter(
            excel_path,
            engine='openpyxl',
            mode='a' if file_exists else 'w',
            if_sheet_exists='overlay' if file_exists else None
        ) as writer:
            if file_exists and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
                df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, header=False, index=False)
                systemLogger.info(f"📝 Excel 追加: sheet={sheet_name}, 起始行={startrow + 1}")
            else:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                systemLogger.info(f"🆕 Excel 新建 sheet: {sheet_name}")

        # 调整列宽 & 数值右对齐（与原版一致）
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]

        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 7

        # D~H 列右对齐（4~8列）
        for row in ws.iter_rows(min_row=2):
            for col_idx in range(4, 9):
                cell = row[col_idx - 1]
                cell.alignment = Alignment(horizontal="right")

        wb.save(excel_path)
        systemLogger.info(f"✅ Excel 保存成功: {excel_path}, 品种={product_prefix}, 日期={stats.get('trading_day')}")
        return True

    except Exception as e:
        systemLogger.exception(f"❌ 保存 Excel 失败: 品种={product_prefix}, 日期={stats.get('trading_day')}, 错误={e}")
        return False

def _parse_contract_yyyymm(instrument_id: str) -> Optional[int]:
    """
    从合约代码尾部解析 YYMM -> 返回 YYYYMM（int）。如 'cu2509' -> 202509。
    不可解析返回 None。
    """
    if not isinstance(instrument_id, str):
        return None
    m = re.search(r'(\d{3,4})$', instrument_id.strip())
    if not m:
        return None
    tail = m.group(1)
    if len(tail) == 4:         # 正常 YYMM
        yy = int(tail[:2]); mm = int(tail[2:])
        yyyy = 2000 + yy
    else:                      # 兜底 YMM
        yy = int(tail[:1]); mm = int(tail[1:])
        yyyy = 2000 + yy
    if 1 <= mm <= 12:
        return yyyy * 100 + mm
    return None

def _is_main_near(main_contract: str, sub_contract: str) -> tuple[bool, str, str, int]:
    """
    判断主力是否为近月。
    返回: (main_is_near, near_contract, far_contract, sign)
      - sign=+1 表示 main 是近月；sign=-1 表示 main 是远月（需展示取反）。
    """
    ym_main = _parse_contract_yyyymm(main_contract)
    ym_sub  = _parse_contract_yyyymm(sub_contract)
    # 无法解析时，默认 main 为近月，sign=+1
    if ym_main is None or ym_sub is None:
        return True, main_contract, sub_contract, +1
    if ym_main <= ym_sub:
        return True, main_contract, sub_contract, +1
    else:
        return False, sub_contract, main_contract, -1

def plot_spread_summary_dashboard(
    product_prefix: str,
    main_contract: str,
    sub_contract: str,
    spread: pd.Series,
    spread_diff: pd.Series,
    exch_times: pd.Series,
    stats: dict,
    save_path: str = None
):
    """
    生成主力-次主力对冲分析仪表盘图表（展示阶段统一为：价差 = 近月 - 远月）。
    """
    try:
        systemLogger.info(f"📊 开始生成仪表盘图表: 品种={product_prefix}, 日期={stats.get('trading_day')}")

        # —— 判定 main/sub 近远月关系（只影响展示，不改原数据）——
        main_is_near, near_c, far_c, sign = _is_main_near(main_contract, sub_contract)
        if main_is_near:
            systemLogger.info(f"➡️ 展示方向: {near_c} - {far_c}（主力为近月，保持原方向）")
        else:
            systemLogger.info(f"🔄 展示方向调整为 近月-远月: {near_c} - {far_c}（主力为远月，展示时取反）")

        # 用于展示的价差序列/价差变化：主力为远月时取反
        spread_to_plot = spread * sign
        spread_diff_to_plot = spread_diff * sign

        # 用于表格“价差”行的总变化：主力为远月时取反
        spread_total_change_disp = (stats.get('spread_total_change', 0) or 0) * sign

        # ——— 布局 ———
        fig = plt.figure(figsize=(13, 7))
        gs = gridspec.GridSpec(3, 2, height_ratios=[0.2, 1.2, 1.6], width_ratios=[0.7, 1.3])

        # ===== 顶部标题 =====
        ax_title = plt.subplot(gs[0, :])
        ax_title.text(0.5, 0.5,
                      f"{stats.get('trading_day')} 品种: {product_prefix.upper()}",
                      fontsize=20, ha='center', va='center', weight='bold')
        ax_title.axis('off')

        # ===== 表格区（价差行按近月-远月展示）=====
        ax_table = plt.subplot(gs[1, :])
        col_labels = ["", "合约", "成交量", "月均成交量", "价格变化量", "波动率", "与价差相关性", "涨跌停"]
        table_data = [
            ["主力合约", main_contract,
             f"{stats.get('vol_main', 0):,.0f}", f"{stats.get('main_avg_month_volume', 0):,.0f}",
             f"{stats.get('leg1_total_change', 0):.2f}", f"{stats.get('leg1_volatility', 0):.4f}",
             f"{stats.get('corr_spread_leg1', 0):.3f}", "是" if stats.get('main_limit_flag') else "否"],

            ["次主力合约", sub_contract,
             f"{stats.get('vol_sub', 0):,.0f}", f"{stats.get('sub_avg_month_volume', 0):,.0f}",
             f"{stats.get('leg2_total_change', 0):.2f}", f"{stats.get('leg2_volatility', 0):.4f}",
             f"{stats.get('corr_spread_leg2', 0):.3f}", "是" if stats.get('sub_limit_flag') else "否"],

            # 价差一行：展示为 近月-远月，并用 sign 修正总变化的符号
            ["价差", f"{near_c}-{far_c}",
             "---", "---",
             f"{spread_total_change_disp:.2f}", f"{stats.get('spread_volatility', 0):.4f}",
             f"{stats.get('corr_leg1_leg2', 0):.3f}", "---"]
        ]

        # 保存表格数据
        save_spread_data_to_excel(
            product_prefix=product_prefix,
            main_contract=main_contract,
            sub_contract=sub_contract,
            stats=stats,  # 原 stats 不改；保存仍按原口径
            excel_path=os.path.join("../output/spread_plots/spread_analysis_summary.xlsx")
        )

        table = ax_table.table(cellText=table_data, colLabels=col_labels, loc='center', cellLoc='center')
        table.auto_set_font_size(False)
        for (row, col), cell in table.get_celld().items():
            if row > 0:
                cell.set_fontsize(12)
        table.scale(1, 2)
        ax_table.axis('off')

        # ===== Spread 分布直方图（展示为近月-远月）=====
        ax1 = plt.subplot(gs[2, 0])
        tick_size = float(stats.get("tick_size", 0.1))
        min_spread = np.floor(spread_to_plot.min() / tick_size) * tick_size
        max_spread = np.ceil(spread_to_plot.max() / tick_size) * tick_size
        bins = np.arange(min_spread, max_spread + tick_size, tick_size)

        counts, bin_edges, patches = ax1.hist(spread_to_plot, bins=bins, color='skyblue', edgecolor='k')
        # 可选：柱顶显示频数
        for c, rect in zip(counts, patches):
            if c > 0:
                ax1.text(rect.get_x() + rect.get_width()/2, rect.get_height(), f"{int(c)}",
                         ha='center', va='bottom', fontsize=8)

        step = max(1, len(bins) // 10)
        xtick_locs = bins[::step]
        ax1.set_xticks(xtick_locs)
        ax1.set_xticklabels([f"{x:.2f}" for x in xtick_locs], rotation=45)
        ax1.set_title("价差分布（近月-远月）")
        ax1.set_xlabel("价差")
        ax1.set_ylabel("频数")
        ax1.grid(True, linestyle='--', alpha=0.5)

        # ===== Spread Diff 趋势图（展示为近月-远月）=====
        ax2 = plt.subplot(gs[2, 1])
        exch_times = pd.to_datetime(exch_times)
        group_size = 10
        x_grouped = list(range(0, len(spread_diff_to_plot), group_size))
        y_grouped = [spread_diff_to_plot.iloc[i:i + group_size].sum() for i in x_grouped]

        ax2.bar(x_grouped, y_grouped, width=group_size, color='skyblue', edgecolor='skyblue', align='edge')
        step2 = max(len(exch_times) // 12, 1)
        tick_locs2 = list(range(0, len(exch_times), step2))
        tick_labels2 = [exch_times[i].strftime("%H:%M") for i in tick_locs2]
        ax2.set_xticks(tick_locs2)
        ax2.set_xticklabels(tick_labels2, rotation=45, ha='right')
        ax2.set_xlim(0, len(spread_diff_to_plot))
        ax2.set_title("价差变化趋势（每10分钟聚合，近月-远月）")
        ax2.set_xlabel("时间")
        ax2.set_ylabel("变化量")
        ax2.grid(True, linestyle='--', alpha=0.5)

        # ===== 涨跌停点标注（沿用原时间坐标）=====
        if "limit_timestamps" in stats and stats["limit_timestamps"]:
            limit_ts = pd.to_datetime(stats["limit_timestamps"])
            limit_indices = [i for i, t in enumerate(exch_times) if t in set(limit_ts)]
            limit_values = [spread_diff_to_plot.iloc[i] for i in limit_indices]
            ax2.scatter(limit_indices, limit_values, color='red', marker='o', s=40, label='涨跌停点')
            ax2.legend(loc='upper right')

        # ——— 输出 ———
        plt.tight_layout(h_pad=2)
        if save_path:
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            plt.savefig(save_path, dpi=150)
            systemLogger.info(f"✅ 仪表盘图表保存成功: {save_path}")
        else:
            plt.show()

        plt.close()

    except Exception as e:
        systemLogger.exception(f"❌ 仪表盘图表生成失败: 品种={product_prefix}, 错误={e}")


def analyze_product(product_prefix: str, tradingday, timescale_crud: TimescaleCRUD, output_root="../output"):
    try:
        systemLogger.info(f"开始分析: 品种={product_prefix}, 交易日={tradingday}")
        tradingday = normalize_trading_day(tradingday)

        # 1. 加载数据
        main_df, sub_df, main_info, sub_info = load_main_sub_data(timescale_crud, tradingday, product_prefix)
        systemLogger.info(f"数据加载完成: 品种={product_prefix}, 合约数=主力{len(main_df)}, 次主力{len(sub_df)}")

        # 获取合约的最小变动单位
        main_contract = main_df.iloc[0]["instrument_id"]
        sub_contract = sub_df.iloc[0]["instrument_id"]
        tick_size = get_tick_size(main_contract, timescale_crud)

        start_time = time.time()
        # 获取最近一月的平均成交量
        main_avg_month_volume = get_avg_volume(timescale_crud, main_contract, tradingday, days=30)
        end_time = time.time()
        systemLogger.info(f"·······主力获取平均成交量耗时: {end_time - start_time}")
        start_time = time.time()
        sub_avg_month_volume = get_avg_volume(timescale_crud, sub_contract, tradingday, days=30)
        end_time = time.time()
        systemLogger.info(f"·········此主力获取平均成交量耗时: {end_time - start_time}")

        # 当日成交量统计
        trade_date = datetime.strptime(tradingday, "%Y-%m-%d").date()
        main_volume = main_df[main_df["trading_day"] == trade_date]["volume"].sum()
        sub_volume = sub_df[sub_df["trading_day"] == trade_date]["volume"].sum()
        systemLogger.info(f"成交量统计: 主力={main_volume}, 次主力={sub_volume}")

        if main_volume < 10000 and sub_volume < 10000:
            systemLogger.warning(f"成交量不足跳过分析: 品种={product_prefix}, 交易日={tradingday}")
            return

        # 净化字段 + 去除开收盘噪声
        main_df = main_df[["instrument_id", "trading_day", "time", "open_price", "high_price", "low_price", "close_price", "volume", "turnover", "total_cnt", "bid_price0", "bid_vol0", "ask_price0", "ask_vol0", "high_limited", "low_limited"]]
        sub_df = sub_df[main_df.columns]
        main_df = remove_open_close_noise(main_df)
        sub_df = remove_open_close_noise(sub_df)

        # 对齐数据
        aligned_df = align_main_sub_minute(main_df, sub_df, tick_size)
        aligned_df, limit_timestamps = filter_spread_with_limit(aligned_df)

        spread = aligned_df['spread']
        spread_diff = spread.diff()
        spread_diff.index = aligned_df['time']

        # 计算单边中间价变化
        leg1_diff = compute_leg_return_series(main_df)
        leg2_diff = compute_leg_return_series(sub_df)

        # 统计指标
        stats = {
            "spread_volatility": compute_spread_volatility(spread_diff),
            "leg1_volatility": compute_leg_volatility(leg1_diff),
            "leg2_volatility": compute_leg_volatility(leg2_diff),
            "vol_main": main_volume,
            "vol_sub": sub_volume,
            "leg1_total_change": leg1_diff.dropna().sum(),
            "leg2_total_change": leg2_diff.dropna().sum(),
            "spread_total_change": spread_diff.dropna().sum(),
            "main_limit_flag": aligned_df["main_limited_flag"].any(),
            "sub_limit_flag": aligned_df["sub_limited_flag"].any(),
            "tick_size": tick_size,
            "limit_timestamps": limit_timestamps,
            "main_avg_month_volume": main_avg_month_volume,
            "sub_avg_month_volume": sub_avg_month_volume,
            "trading_day": tradingday
        }

        corrs = compute_spread_leg_correlation(spread_diff, leg1_diff, leg2_diff)
        stats.update({
            "corr_spread_leg1": corrs["spread_vs_leg1"],
            "corr_spread_leg2": corrs["spread_vs_leg2"],
            "corr_leg1_leg2": corrs["leg1_vs_leg2"],
        })
        systemLogger.info(f"指标计算完成: 品种={product_prefix}, 交易日={tradingday}, 统计数据={stats}")

        # 生成图表
        plot_dir = os.path.join(output_root, "spread_plots")
        os.makedirs(plot_dir, exist_ok=True)
        dashboard_path = os.path.join(plot_dir, f"{product_prefix}_仪表盘.png")
        plot_spread_summary_dashboard(
            product_prefix=product_prefix,
            main_contract=main_contract,
            sub_contract=sub_contract,
            spread=spread,
            spread_diff=spread_diff,
            exch_times=aligned_df["time"],
            stats=stats,
            save_path=dashboard_path
        )
        systemLogger.info(f"图表保存成功: {dashboard_path}")
        systemLogger.info(f"✅ 分析完成: 品种={product_prefix}, 交易日={tradingday}")

    except Exception as e:
        systemLogger.exception(f"❌ 分析失败: 品种={product_prefix}, 交易日={tradingday}, 错误: {e}")


def clean_volume(vol):
    """清洗成交量数据"""
    try:
        return int(str(vol).split("(")[0].replace(",", ""))
    except:
        return None

def plot_metric_comparison_bar(
    df_main,
    df_sub,
    df_spread,
    sheet_name,
    metric_key,
    metric_format_map=None,
    show_labels=True,
    label_threshold_ratio=0.01,   # 小柱子不显示
    delta_ratio_threshold=0.25,   # 同组过于接近的只显示一个
    contract_col_candidates=("合约", "合约代码", "instrument_id"),  # 兼容不同列名
    save_dir="../output/spread_plots"  # 文件保存目录
):
    """
    比较主力、次主力、价差三者在同一指标下的柱状图。

    颜色策略：
      - 主力：以 #ff7f0e 为基色（橙色），不同主力合约仅做轻微色阶变化（最多3档）
      - 次主力：以 #1f77b4 为基色（蓝色），不同次主力合约仅做轻微色阶变化（最多3档）
      - 价差：固定绿色 #2ca02c
    Legend 展示：主力/次主力的「合约代码 → 颜色」映射。
    """

    # ===== 默认格式映射 =====
    if metric_format_map is None:
        metric_format_map = {
            "成交量": "int",
            "价格变化量": "float",
            "波动率": "float",
            "与价差相关性": "float"
        }

    # ===== 工具函数：轻微色阶（最多3档）=====
    def generate_shades(base_hex: str, n: int, lightness_step: float = 0.17):
        """
        基于基色，生成最多 n 个轻微亮度变化的色阶（-step, 0, +step）。
        色差很小，便于同类区分，不会跨度太大。
        """
        n = max(1, min(10, n))  # 限制到 1~5
        base_hex = base_hex.lstrip("#")
        r, g, b = [int(base_hex[i:i+2], 16)/255.0 for i in (0, 2, 4)]
        h, l, s = colorsys.rgb_to_hls(r, g, b)

        # 为了“微调”，顺序优先给基色，再略暗、再略亮（或可相反）
        deltas = [0.0, -lightness_step, lightness_step][:n]

        shades = []
        for d in deltas:
            l_new = max(0.0, min(1.0, l + d))
            r_new, g_new, b_new = colorsys.hls_to_rgb(h, l_new, s)
            shades.append(f"#{int(r_new*255):02x}{int(g_new*255):02x}{int(b_new*255):02x}")
        return shades

    # ===== 预处理数据 =====
    df_main = df_main.sort_values("交易日").copy()
    df_sub = df_sub.sort_values("交易日").copy()
    df_spread = df_spread.sort_values("交易日").copy()

    def pick_contract_col(df):
        for c in contract_col_candidates:
            if c in df.columns:
                return c
        return None

    main_contract_col = pick_contract_col(df_main)
    sub_contract_col  = pick_contract_col(df_sub)

    dates = pd.to_datetime(df_main["交易日"]).dt.strftime("%Y-%m-%d").tolist()
    x = list(range(len(dates)))

    def extract_values(df):
        vals = []
        for v in df[metric_key]:
            try:
                vals.append(float(str(v).split("(")[0]))
            except Exception:
                vals.append(None)
        return vals

    y_main = extract_values(df_main)
    y_sub = extract_values(df_sub)
    y_spread = extract_values(df_spread)

    # ===== 颜色映射：按合约首次出现顺序分配轻微色阶 =====
    def build_color_by_contract(df, contract_col, base_hex):
        """
        返回：
          colors: 与 df 行一一对应的颜色
          cmap  : {contract -> color}（legend 用）
        """
        if contract_col is None or contract_col not in df.columns:
            return [base_hex] * len(df), {}

        seq = pd.Series(df[contract_col]).fillna("UNKNOWN").astype(str).tolist()

        # 按出现顺序去重
        unique_contracts = []
        seen = set()
        for c in seq:
            if c not in seen:
                seen.add(c)
                unique_contracts.append(c)

        # 生成轻微色阶（最多3档）
        shades = generate_shades(base_hex, len(unique_contracts))

        # 由于最多3档，如果 unique_contracts > 3，就循环使用（你说通常最多3次变更）
        color_map = {c: shades[i % len(shades)] for i, c in enumerate(unique_contracts)}
        colors = [color_map[c] for c in seq]
        return colors, color_map

    main_colors, main_color_map = build_color_by_contract(df_main, main_contract_col, "#ff7f0e")  # 橙
    sub_colors, sub_color_map   = build_color_by_contract(df_sub,  sub_contract_col,  "#1f77b4")  # 蓝
    spread_colors = ["#2ca02c"] * len(df_spread)  # 绿

    # ===== 绘图 =====
    bar_width = 0.25
    plt.figure(figsize=(9, 4))

    x_main = [i - bar_width for i in x]
    x_sub = x
    x_spread = [i + bar_width for i in x]

    bars1 = plt.bar(x_main, y_main, width=bar_width, label="主力", color=main_colors)
    bars2 = plt.bar(x_sub, y_sub, width=bar_width, label="次主力", color=sub_colors)
    bars3 = plt.bar(x_spread, y_spread, width=bar_width, label="价差", color=spread_colors)

    # 动态 y 轴范围
    all_y = [v for v in y_main + y_sub + y_spread if v is not None]
    if all_y:
        y_min, y_max = min(all_y), max(all_y)
        if y_max - y_min < 1e-6:
            plt.ylim(y_min - 0.01, y_max + 0.01)
        else:
            pad = (y_max - y_min) * 0.2
            plt.ylim(y_min - pad, y_max + pad)

    # ===== 标签：每列最多显示一个（最大/差异足够） =====
    if show_labels:
        fmt_type = metric_format_map.get(metric_key, "float")
        for i in range(len(x)):
            bars_group = []
            values_group = []
            for bar_list, y_list in zip([bars1, bars2, bars3], [y_main, y_sub, y_spread]):
                bar = bar_list[i]
                height = bar.get_height()
                if pd.notnull(height):
                    bars_group.append(bar)
                    values_group.append(height)

            if not values_group:
                continue

            # 绝对值排序
            sorted_items = sorted(zip(bars_group, values_group), key=lambda kv: abs(kv[1]), reverse=True)
            max_val = abs(sorted_items[0][1])

            for j, (bar, val) in enumerate(sorted_items):
                if abs(val) < abs(max(all_y)) * label_threshold_ratio:
                    continue
                if j > 0:
                    diff_ratio = abs(max_val - abs(val)) / max_val
                    if diff_ratio < delta_ratio_threshold:
                        continue

                # 数值格式：≥10不保留小数；或显式 int 指标
                if fmt_type == "int" or abs(val) >= 10:
                    label = f"{val:.0f}"
                else:
                    label = f"{val:.2f}"

                offset = val * 0.05 + 0.01
                y_pos = val + offset if val >= 0 else val - offset
                va = 'bottom' if val >= 0 else 'top'

                plt.text(
                    bar.get_x() + bar.get_width() / 2,
                    y_pos,
                    label,
                    ha='center',
                    va=va,
                    fontsize=5
                )

    # ===== Legend：主力/次主力合约映射 + 价差 =====
    legend_handles = []
    if main_color_map:
        legend_handles.append(Patch(facecolor="none", edgecolor="none", label="— 主力合约 —"))
        for c, col in main_color_map.items():
            legend_handles.append(Patch(facecolor=col, edgecolor="none", label=f"主力: {c}"))
    if sub_color_map:
        legend_handles.append(Patch(facecolor="none", edgecolor="none", label="— 次主力合约 —"))
        for c, col in sub_color_map.items():
            legend_handles.append(Patch(facecolor=col, edgecolor="none", label=f"次主力: {c}"))
    legend_handles.append(Patch(facecolor="#2ca02c", edgecolor="none", label="价差"))

    plt.xticks(ticks=x, labels=dates, rotation=45, fontsize=5, ha='right')
    plt.xlabel("交易日", fontsize=6)
    plt.ylabel(metric_key, fontsize=6)
    plt.title(f"{sheet_name} - {metric_key} 对比", fontsize=6)
    plt.grid(True, axis='y', linestyle="--", alpha=0.5)
    plt.legend(handles=legend_handles, fontsize=5, ncols=1, loc="best")

    # ===== 保存 =====
    plt.tight_layout(pad=0.5)
    plt.subplots_adjust(bottom=0.25)

    os.makedirs(save_dir, exist_ok=True)
    safe_metric = str(metric_key).replace("/", "_").replace("\\", "_")
    out_path = os.path.join(save_dir, f"{sheet_name.lower()}_{safe_metric}.png")

    img_buffer = BytesIO()
    plt.savefig(img_buffer, format="png", dpi=200, bbox_inches="tight")
    plt.savefig(out_path, format="png", dpi=200, bbox_inches="tight")
    img_buffer.seek(0)
    plt.close()

    return Image(img_buffer)



def add_bar_charts_to_excel(excel_path, output_excel):
    systemLogger.info("📊 开始处理带图表的Excel生成任务")
    if not os.path.exists(excel_path):
        systemLogger.error(f"❌ 源文件不存在: {excel_path}")
        return

    try:
        with pd.ExcelWriter(output_excel, engine="openpyxl", mode="w") as writer:
            with pd.ExcelFile(excel_path) as xls:
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet)
                    df.to_excel(writer, sheet_name=sheet, index=False)
        wb = load_workbook(output_excel)

        with pd.ExcelFile(excel_path) as xls:
            for sheet_name in xls.sheet_names:
                systemLogger.info(f"📄 处理sheet: {sheet_name}")
                df = pd.read_excel(xls, sheet_name=sheet_name)
                ws = wb[sheet_name]

                df["交易日"] = pd.to_datetime(df["交易日"])
                df_main = df[df["合约类型"] == "主力合约"].copy()
                df_sub = df[df["合约类型"] == "次主力合约"].copy()
                df_spread = df[df["合约类型"] == "价差"].copy()
                df_main["成交量"] = df_main["成交量"].apply(clean_volume)
                df_sub["成交量"] = df_sub["成交量"].apply(clean_volume)

                current_row = 1
                metrics = ["成交量", "价格变化量", "波动率", "与价差相关性"]
                for metric in metrics:
                    img = plot_metric_comparison_bar(df_main, df_sub, df_spread, sheet_name, metric)
                    ws.add_image(img, anchor=f"A{current_row}")
                    current_row += 40  # 图高些可以适当增加间隔

                for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
                    max_length = 0
                    for cell in col:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 5

                for row in ws.iter_rows(min_row=2):
                    for col_idx in range(4, 9):
                        cell = row[col_idx - 1]
                        cell.alignment = Alignment(horizontal="right")

        wb.save(output_excel)
        systemLogger.info(f"✅ 图表已成功插入并保存至: {output_excel}")

    except Exception as e:
        systemLogger.exception(f"❌ 图表生成失败: {e}")




def generate_html_with_anchor(image_dir: str) -> Tuple[str, list[str]]:
    """
    生成含目录锚点跳转的 HTML，用于邮件中内嵌图像，并兼容 PC 企业微信。
    每个品种下图表展示顺序固定为：dashboard → 成交量 → 价格变化量 → 波动率 → 与价差相关性
    """
    image_files = sorted([
        f for f in os.listdir(image_dir)
        if f.lower().endswith((".png", ".jpg", ".jpeg"))
    ])

    # 图名排序优先级
    metric_order = ["仪表盘", "成交量", "价格变化量", "波动率", "与价差相关性"]

    # 按品种分组
    product_map = defaultdict(list)
    for filename in image_files:
        product = filename.split("_")[0].upper()
        product_map[product].append(filename)

    html = ["<html><head><meta charset='utf-8'></head><body id='top'>"]

    # === 更新提示 ===
    html.append(f"""
    <div style="
        background-color:#fff5f5;
        border-left:5px solid #d9534f;
        padding:10px 15px;
        margin-bottom:12px;
        font-size:14px;
        line-height:1.6;
    ">
        <span style="font-size:16px;">🔄</span>
        <strong>更新日期：</strong>
        <span style="color:#c9302c;">2025-08-11</span>

        <br>
        <span style="font-size:16px;">📝</span>
        <strong>更新内容：</strong>
        ① 价差由原本的 <strong>主力-次主力</strong> 更改为 <strong>近月-远月</strong>；
        ② 原本主力、次主力成交量 <strong>任意</strong> 低于 <code>10000</code> 就不统计，
        更改为主力、次主力成交量 <strong>都</strong> 低于 <code>10000</code> 不统计。
        ③ 月趋势图中主力与次主力合约，通过修改不同色阶区分换月情况。
        <br>

        <span style="font-size:16px;">💡</span>
        <strong>使用建议：</strong>
        建议使用 <strong>企业微信手机客户端</strong> 或 <strong>PC 客户端</strong> 浏览，
        键盘<code>Home</code>键或点击返回目录按钮可一键回到目录，提升阅读与跳转效率。

        <br>
        <span style="font-size:16px;">💬</span>
        <strong>反馈建议：</strong>
        如果您对本报告有修改意见或建议，请回复邮件至
        <a href="mailto:chenguangtao@aifanzhellc.com" style="color:#1a73e8; text-decoration:none;">
            chenguangtao@aifanzhellc.com
        </a>
        <br>

    </div>
    """)

    # === 顶部目录导航 ===
    html.append("""
    <h2 id='目录' style='font-size:16px;'>📌 品种目录</h2>
    <ul style='font-size:14px; list-style: none; padding-left: 0;'>
    """)
    for product in sorted(product_map):
        html.append(f"""
            <li style="margin-bottom: 18px;">
                👉 <a href="#{product}" style="color:#1a73e8; text-decoration: none;">{product}</a>
                <hr style="border: none; border-top: 1px solid #eee; margin: 4px 0 4px 20px;">
            </li>
        """)
    html.append("</ul><hr>")

    # === 主体内容 ===
    img_paths = []
    img_index = 1
    for product in sorted(product_map):
        html.append(f'<h3 id="{product}" style="font-size:14px; border-bottom:1px solid #ccc;">{product}</h3>')

        # 按图名关键字进行排序
        sorted_files = sorted(
            product_map[product],
            key=lambda x: next((i for i, metric in enumerate(metric_order) if metric in x), len(metric_order))
        )

        for filename in sorted_files:
            file_name_no_ext = os.path.splitext(filename)[0]
            html.append(f'<p><b>{file_name_no_ext}</b><br><img></p>')
            img_paths.append(os.path.join(image_dir, filename))
            img_index += 1
            html.append('<p style="text-align:left;"><a href="#目录" style="font-size:15px;">🔝 返回目录</a></p>')
            html.append('<hr>')

    html.append("</body></html>")
    return "\n".join(html), img_paths



def clear_folder(folder_path: str):
    shutil.rmtree(folder_path)
    os.makedirs(folder_path, exist_ok=True)


def Usage():
    print('Usage:###################################')
    print('#Usage    : python tradeAssistantSpread.py <startDay> <endDay>')
    print('#param1   : <startDay>   eg: 20250806')
    print('#param2   : <endDay>     eg: 20250806')
    print('##########################################')


# if __name__ == "__main__":

#     if len(sys.argv) != 3:
#         Usage()
#         exit(1)

#     start_day = str(sys.argv[1]).strip()
#     end_day = str(sys.argv[2]).strip()

#     start_day = 20250728
#     end_day = 20250828
#     # 清空"../output/spread_plots/”该文件夹下的内容
#     clear_folder("../output/spread_plots/")

#     excel_path = os.path.join("../output/spread_plots/spread_analysis_summary.xlsx")
#     output_excel = os.path.join("../output/spread_plots/spread_analysis_with_bar_charts_4.xlsx")

#     timescale_crud = TimescaleCRUD(DATABASE_URL)
#     trading_days = get_trading_days_in_range(start_day, end_day, timescale_crud)

#     for tradingday in trading_days:
#         for product in ["cu", "zn", "al", "ni", "i", "au", "IH", "IF", "IC", "IM"]:
#                 analyze_product(product, tradingday, timescale_crud, output_root="../output")


#     # # 处理表格信息
#     add_bar_charts_to_excel(excel_path, output_excel)

#     # 邮件发送
#     html_text, image_paths = generate_html_with_anchor("../output/spread_plots/")
#     mail = Mail(
#         receivers=RECEIVERS,
#         host=SMTP_SERVER,
#         port=SMTP_PORT,
#         user=EMAIL_ADDRESS,
#         password=EMAIL_PASSWORD
#     )

#     mail.add_tittle("价差统计图合集", from_nickname="策略研究助手")

#     # 添加图片
#     for img_path in image_paths:
#         mail.add_img(img_path)

#     # 添加 HTML 内容（含跳转目录）
#     mail.add_html_spread(html_text)
#     # 添加 Excel 附件
#     mail.add_excel_attachment(excel_path)
#     mail.send()
